import fitz  # Импортируем PyMuPDF
from openpyxl import Workbook, load_workbook
import os
import re
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def read_highlighted_texts_and_comments(pdf_path):
    # print(f"Попытка открыть файл: {pdf_path}")  # Выводим имя файла перед открытием
    doc = fitz.open(pdf_path)
    highlighted_texts = []
    comments = []

    for page in doc:
        annotations = page.annots()
        if annotations:
            for annot in annotations:
                if annot.type[0] == 8:  # Highlight annotation
                    # Уменьшаем прямоугольник аннотации для уточнения области поиска
                    rect = annot.rect
                    adjusted_rect = fitz.Rect(rect.x0, rect.y0 + (rect.y1 - rect.y0) * 0.3, rect.x1, rect.y1 - (rect.y1 - rect.y0) * 0.3)
                    words = page.get_text("words")
                    highlighted_words = [w[4] for w in words if fitz.Rect(w[:4]).intersects(adjusted_rect)]
                    if highlighted_words:  # Если нашлись слова внутри аннотации
                        highlighted_text = " ".join(highlighted_words)
                        highlighted_texts.append(highlighted_text)
                        # Чтение комментария, если он есть
                        comment = annot.info['content'] if 'content' in annot.info else ''
                        print(f"Найден текст: {highlighted_text}, комментарий: {comment}")
                        comments.append(comment)

    doc.close()
    return highlighted_texts, comments


def calculate_km(sum_value, kokku_value):
    try:
        sum_float = parse_sum_to_float(sum_value) if sum_value else 0.0
        kokku_float = parse_sum_to_float(kokku_value) if kokku_value else sum_float
        return kokku_float - sum_float
    except ValueError:
        return 0.00  # или 0, если предпочтительнее возвращать 0 при ошибках преобразования


def process_highlighted_texts_and_comments(highlighted_texts, comments, predefined_keys):
    output_dict = {key: None for key in predefined_keys}
    unmatched_comments = []  # Для комментариев, не соответствующих ключам

    for text, comment in zip(highlighted_texts, comments):
        if comment in predefined_keys:
            if output_dict[comment] is None:
                output_dict[comment] = text
        else:
            unmatched_comments.append(comment)  # Собираем несоответствующие комментарии

    #print(output_dict)

    output_dict['km'] = calculate_km(output_dict.get('sum'), output_dict.get('kokku'))

    if output_dict["sum"] is not None:
        output_dict["sum"] = parse_sum_to_float(output_dict["sum"])
    if output_dict["kokku"] is not None:
        output_dict["kokku"] = parse_sum_to_float(output_dict["kokku"])

    if output_dict["date"] is not None:
        output_dict["date"] = format_date(output_dict["date"])
    if output_dict["date2"] is not None:
        output_dict["date2"] = format_date(output_dict["date2"])

    # Специальная обработка для некоторых ключей
    if output_dict["kokku"] is None:
        output_dict["kokku"] = output_dict["sum"]

    if output_dict["date2"] is None and output_dict["date"] is not None:
        output_dict["date2"] = output_dict["date"]
    if output_dict["cur"] is None:
        output_dict["cur"] = "EUR"

    if output_dict["our"] is None:
        output_dict["our"] = "_"

    #print(output_dict)

    return output_dict, unmatched_comments


def reverse_date_format(date_str):
    # Разделяем исходную строку даты на день, месяц и год
    day, month, year = date_str.split(".")
    # Извлекаем последние две цифры года и переставляем элементы в нужном порядке
    reversed_date = f"{year[-2:]}{month}{day}"
    return reversed_date


def parse_sum_to_float(value):
    # Извлекаем числа, пробелы (для тысяч), точку и запятую (как десятичные разделители)
    numeric_value = re.sub(r"[^\d,. ]", "", value)
    # Удаляем пробелы
    numeric_value_without_spaces = numeric_value.replace(" ", "")
    # Определяем, используется ли точка или запятая в качестве десятичного разделителя
    # Если и точка, и запятая присутствуют, предполагаем, что запятая используется для разделения тысяч, если запятая стоит раньше
    if ',' in numeric_value_without_spaces and '.' in numeric_value_without_spaces:
        if numeric_value_without_spaces.index(',') < numeric_value_without_spaces.index('.'):
            # Запятая для разделения тысяч, удаляем запятые
            numeric_value_without_spaces = numeric_value_without_spaces.replace(",", "")
        else:
            # Точка для разделения тысяч, удаляем точки и заменяем запятую на точку
            numeric_value_without_spaces = numeric_value_without_spaces.replace(".", "").replace(",", ".")
    elif ',' in numeric_value_without_spaces:
        # Запятая в качестве десятичного разделителя
        numeric_value_without_spaces = numeric_value_without_spaces.replace(",", ".")
    # В этот момент все запятые уже заменены на точки, если они использовались как десятичные разделители
    return float(numeric_value_without_spaces)


def format_date(date_str):
    month_mapping = {
        "01": ["jaanuar", "january"],
        "02": ["veebruar", "february"],
        "03": ["märts", "march"],
        "04": ["aprill", "april"],
        "05": ["mai", "may"],
        "06": ["juuni", "june"],
        "07": ["juuli", "july"],
        "08": ["august", "august"],
        "09": ["september", "september"],
        "10": ["oktoober", "october"],
        "11": ["november", "november"],
        "12": ["detsember", "december"],
    }

    # Преобразование названия месяца в его числовой эквивалент
    date_str_processed = date_str.lower()
    for month_number, month_names in month_mapping.items():
        for month_name in month_names:
            if month_name in date_str_processed:
                replace_with = "." + month_number + "."  # Добавляем точку после числа месяца
                date_str_processed = re.sub(month_name, replace_with, date_str_processed, flags=re.IGNORECASE)
                date_str_processed = date_str_processed.replace("..", ".")  # Удаляем двойные точки, если они есть
                break

    # Удаление лишних пробелов
    date_str_processed = re.sub(r'\s+', '', date_str_processed).strip()

    # Пытаемся разобрать обработанную строку даты
    formatted_date = parse_date_from_common_formats(date_str_processed)
    if formatted_date:
        return formatted_date

    # Если специфическая обработка не привела к успеху, пытаемся разобрать исходную строку
    return parse_date_from_common_formats(date_str)


def parse_date_from_common_formats(date_str):
    """
    Пытается разобрать дату из строки, используя список стандартных форматов.
    Возвращает форматированную строку даты "%d.%m.%Y" или None, если разбор не удался.
    """
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y", "%d-%m-%y", "%d/%m/%y", "%d.%m.%y", "%Y-%m-%d", "%y-%m-%d"):
        try:
            parsed_date = datetime.strptime(date_str, fmt)
            return parsed_date.strftime("%d.%m.%Y")
        except ValueError:
            continue
    return None


def auto_adjust_column_width(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column  # Получаем номер столбца
        for cell in col:
            try:  # Необходимо для обработки текста и None-значений
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Добавляем немного дополнительного пространства
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width


def add_table_to_sheet(sheet, table_name="Table1"):
    # Проверяем, существует ли уже таблица с таким именем
    existing_tables = [table.displayName for table in sheet.tables.values()]
    unique_name = table_name
    counter = 1
    while unique_name in existing_tables:
        unique_name = f"{table_name}_{counter}"
        counter += 1

    # Определяем диапазон таблицы
    max_column = sheet.max_column
    max_row = sheet.max_row
    ref = f"A1:{get_column_letter(max_column)}{max_row}"

    # Создаем объект таблицы
    table = Table(displayName=unique_name, ref=ref)

    # Создаем и применяем стиль таблицы
    style = TableStyleInfo(name="TableStyleLight14", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style

    # Добавляем таблицу на лист
    sheet.add_table(table)

    # После добавления таблицы на лист, применяем форматирование к столбцам
    column_names_to_format = ["sum", "km", "kokku"]
    format_string = '#,##0.00'
    apply_format_to_columns(sheet, column_names_to_format, format_string)



def apply_format_to_columns(sheet, column_names_to_format, format_string):
    max_row = sheet.max_row
    # Получаем заголовки столбцов
    headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    # Итерируемся по заголовкам для нахождения индексов нужных столбцов
    for index, column_name in enumerate(headers):
        if column_name in column_names_to_format:
            column_letter = get_column_letter(index + 1)  # Преобразуем индекс в букву столбца
            # Применяем формат ко всем ячейкам в столбце
            for row in range(2, max_row + 1):  # Начинаем с 2, чтобы пропустить заголовок
                cell = sheet[f"{column_letter}{row}"]
                cell.number_format = format_string


def write_to_excel(output_dict, excel_file_name, predefined_keys, new_filename):
    # Проверка и инициализация рабочей книги Excel
    try:
        wb = load_workbook(excel_file_name)
    except FileNotFoundError:
        wb = Workbook()
    sheet = wb.active

    if sheet.max_row == 1 and all([cell.value is None for cell in sheet[1]]):
        # Если файл только что создан, добавляем заголовки столбцов
        for idx, key in enumerate(predefined_keys, 1):
            sheet.cell(row=1, column=idx, value=key)
        next_row = 2
    else:
        next_row = sheet.max_row + 1

    for idx, key in enumerate(predefined_keys, 1):
        value = output_dict.get(key, "")
        if isinstance(value, float):
            sheet.cell(row=next_row, column=idx, value=value)
        else:
            sheet.cell(row=next_row, column=idx, value=str(value))

    # Добавление гиперссылки на файл PDF
    hyperlink_column = len(predefined_keys) + 1  # Следующий столбец после последнего ключа
    sheet.cell(row=1, column=hyperlink_column, value="Link")  # Называем столбец "Link"
    hyperlink_formula = f'=HYPERLINK("{new_filename}","{new_filename}")'
    sheet.cell(row=next_row, column=hyperlink_column, value=hyperlink_formula)

    wb.save(excel_file_name)


def initialize_excel_file(excel_file_name, predefined_keys):
    wb = Workbook()
    sheet = wb.active
    # Добавляем названия столбцов в первую строку
    for idx, key in enumerate(predefined_keys, 1):
        sheet.cell(row=1, column=idx, value=key)
    wb.save(excel_file_name)


def rename_pdf_file(original_path, date_str, firma, invoice, is_kres_auto=False):
    date = datetime.strptime(date_str, "%d.%m.%Y")
    reverse_date = date.strftime("%y%m%d")

    # Определение формата нового имени файла в зависимости от условия
    if is_kres_auto:
        new_filename = f"Kres_{reverse_date}_{sanitize_filename(firma)}_arve.pdf"
    else:
        new_filename = f"{reverse_date}_{sanitize_filename(firma)}_{sanitize_filename(invoice)}.pdf"

    new_path = os.path.join(os.path.dirname(original_path), new_filename)

    try:
        os.rename(original_path, new_path)
        print(f"Файл '{original_path}' был переименован в '{new_filename}'")
    except OSError as error:
        print(f"Ошибка при переименовании файла: {error}")
    return new_filename


def sanitize_filename(filename):
    """
    Удаляет недопустимые символы из имени файла и возвращает очищенное имя.
    """
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        filename = filename.replace(char, '')
    # Удаление конечных пробелов и точек из имени файла
    filename = filename.rstrip('. ')
    return filename


def process_all_pdfs_in_folder(folder_path, excel_file_name, predefined_keys):
    initialize_excel_file(excel_file_name, predefined_keys)

    for file in os.listdir(folder_path):
        if file.endswith(".pdf"):
            full_path = os.path.join(folder_path, file)
            print(f"Обрабатываем файл: {file}")
            highlighted_texts, comments = read_highlighted_texts_and_comments(full_path)
            output_dict, unmatched_comments = process_highlighted_texts_and_comments(highlighted_texts, comments, predefined_keys)

            if unmatched_comments:
                print(f"Несоответствующие комментарии: {unmatched_comments}")

            none_keys = [key for key, value in output_dict.items() if value is None]
            if none_keys:
                print(f"Нет данных для ключей: {none_keys}")



            # Интеграция функции переименования
            # Предположим, что `output_dict` содержит ключи 'date', 'firma', и 'invoice'
            date_str = output_dict.get('date')  # Например, '01.01.2024'
            firma = output_dict.get('firma')  # Например, 'YourCompany'
            arve = output_dict.get('arve')  # Например, '12345'

            if date_str and firma and arve:
                # Переименовываем файл
                is_kres_auto = "Kres" in output_dict.get("our", "")
                new_filename = rename_pdf_file(full_path, date_str, firma, arve, is_kres_auto)
            else:
                print(f"*** Не удалось переименовать файл {file}, так как не все данные доступны.")

            # Теперь правильно вызываем функцию записи в Excel
            write_to_excel(output_dict, excel_file_name, predefined_keys, new_filename)

    # Загружаем рабочую книгу и лист
    wb = load_workbook(excel_file_name)
    sheet = wb.active

    # Добавляем умную таблицу
    add_table_to_sheet(sheet)

    # После добавления всех данных вызываем функцию auto_adjust_column_width
    auto_adjust_column_width(sheet)


    # Сохраняем изменения
    wb.save(excel_file_name)



def main():
    folder_path = "/Users/docha/Library/CloudStorage/GoogleDrive-kres.auto79@gmail.com/Мой диск/2024-02"
    predefined_keys = ["firma", "date", "arve", "sum", "km", "kokku", "date2", "cur", "our"]

    # Формирование пути к файлу Excel внутри обрабатываемой папки
    excel_file_name = os.path.join(folder_path, "_all_invoices.xlsx")

    process_all_pdfs_in_folder(folder_path, excel_file_name, predefined_keys)


if __name__ == '__main__':
    main()

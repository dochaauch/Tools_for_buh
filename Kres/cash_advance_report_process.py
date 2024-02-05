import openpyxl
from collections import OrderedDict
import string


def read_excel(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name, data_only=True)
    sheet = workbook[sheet_name]

    # Чтение заголовков из первой и второй строки
    headers_row_1 = [cell.value for cell in sheet[1]]
    headers_row_2 = [cell.value for cell in sheet[2]]

    # Объединение заголовков из первой и второй строки
    headers = [f'{r1}//{r2}' if r1 else r2 for r1, r2 in zip(headers_row_1, headers_row_2)]

    # Чтение данных, начиная с третьей строки
    svod_data = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        row_data = {headers[i]: cell for i, cell in enumerate(row)}
        svod_data.append(row_data)

    return svod_data


def create_new_sheet_with_headers(file_name, new_sheet_name, headers):
    workbook = openpyxl.load_workbook(file_name)

    # Проверка наличия листа и его удаление при необходимости
    if new_sheet_name in workbook.sheetnames:
        del workbook[new_sheet_name]

    # Создание нового листа и добавление заголовков
    new_sheet = workbook.create_sheet(new_sheet_name)
    new_sheet.append(headers)

    # Сохранение изменений в файле
    workbook.save(file_name)


def extract_expence_accs_by_prefix(row, prefix):
    accounts = []
    for key, value in row.items():
        if key.startswith(prefix) and value is not None:
            parts = key.split('//')
            account_number = parts[0]
            account_description = parts[1] if len(parts) > 1 else ''  # Описание счета, если оно есть
            accounts.append((account_number, account_description, value))
    return accounts

def process_sheet_data_for_new_sheet(svod_data, column_mappings, headers, kaibemaks, kaibemaksuta_desc,
                                     recalculate_tax_and_total=True):
    zagruzka_data = []
    document_start_index = 0

    for row in svod_data:
        expense_accounts = extract_expence_accs_by_prefix(row, expense_account_prefix)  # Извлекает счета расходов
        km_o_total_for_row = 0
        total_sum_for_doc = 0

        if not expense_accounts:
            expense_accounts = [('', None)]

        for account, account_desc, sum_value in expense_accounts:
            new_row, km_o_value, summa_total = create_row_for_zagruzka(row, account, account_desc,
                                                          sum_value, kaibemaks, kaibemaksuta_desc, headers,
                                                          row_settings)
            km_o_total_for_row += km_o_value
            total_sum_for_doc += summa_total

            for header, svod_key in column_mappings.items():
                new_row[header] = row.get(svod_key, '')

            zagruzka_data.append(new_row)

        if recalculate_tax_and_total:
            # Пересчитываем сумму НСО
            if km_o_total_for_row != round(row.get('km', 0), 2):
                correct_values(zagruzka_data, document_start_index, km_o_total_for_row, row, 'Сумма НСО',
                               'km', "Старый налог")

            # Пересчитываем общую сумму после корректировки НСО
            recalculated_total_sum = sum(float(r['Сумма с НСО']) for r in zagruzka_data[document_start_index:])

            if round(recalculated_total_sum,2) != round(row.get('summa kokku', 0), 2):
                correct_values(zagruzka_data, document_start_index, recalculated_total_sum, row, 'Сумма с НСО',
                               'summa kokku', "Старая сумма")

        document_start_index = len(zagruzka_data)

    return zagruzka_data


def create_row_for_zagruzka(row, account, account_desc, sum_value, kaibemaks, kaibemaksuta_desc, headers,
                            row_settings):
    new_row = {header: '' for header in headers}
    new_row['Счет'] = account
    new_row['Сумма'] = str(sum_value) if sum_value is not None else ''
    km_doc = row.get('km', 0)

    if account_desc not in kaibemaksuta_desc and sum_value is not None and km_doc != 0:
        km_o_value = round(sum_value * kaibemaks, 2)
        new_row['Ставка НСО'] = f'{int(kaibemaks * 100)}%'
        #new_row['Комментарий'] = f"{row.get('Currency', '')} {row.get('kirjeldus', '')}".strip() \
        #    if row.get('Currency', '') else row.get('kirjeldus', '')

    else:
        km_o_value = 0
        new_row['Ставка НСО'] = 'Нет'
        #new_row['Комментарий'] = account_desc
        #new_row['Комментарий'] = f"{row.get('Currency', '')} {row.get('kirjeldus', '')}".strip() \
        #    if row.get('Currency', '') else row.get('kirjeldus', '')

    currency = row.get('Currency', '')
    kirjeldus = row.get('kirjeldus', '')
    new_row['Комментарий'] = create_unique_comment(currency, kirjeldus)

    summa_total = sum_value + km_o_value

    new_row['Сумма НСО'] = f'{km_o_value:.2f}'
    new_row['Сумма с НСО'] = f'{summa_total:.2f}'

    new_row['Вид документа НСО'] = row_settings['doc_type']
    new_row['Применение НСО'] = row_settings['application_nso']
    new_row['Аналитика по НСО'] = row_settings['analytics_nso']
    new_row['Счет проводки НСО'] = row_settings['account_transaction_nso']
    new_row['Валюта отчета'] = row_settings['currency']
    new_row['Курс валюты отчета'] = row_settings['currency_rate']
    return new_row, km_o_value, summa_total


def create_unique_comment(currency, kirjeldus):
    currency = currency or ''
    kirjeldus = kirjeldus or ''

    # Удаление знаков препинания из каждого слова
    words = f'{currency} {kirjeldus}'.split()
    words = [word.strip(string.punctuation) for word in words]

    unique_words = list(OrderedDict.fromkeys(words))
    return ' '.join(unique_words)


def correct_values(zagruzka_data, document_start_index, total_for_row, row, value_key, total_key, message):
    document_id = row.get('arve nr', '')
    total_value = round(row.get(total_key, 0), 2)

    for correction_row in reversed(zagruzka_data[document_start_index:]):
        existing_value = float(correction_row[value_key]) if correction_row[value_key] else 0
        if existing_value != 0:
            correction = total_value - total_for_row
            corrected_value = existing_value + correction
            correction_row[value_key] = f"{corrected_value:.2f}"
            print(f"Документ {document_id}, {row.get('firma', )}: {message} {existing_value:.2f},"
                  f" новое значение {corrected_value:.2f}")
            break


def write_data_to_sheet(file_name, sheet_name, data, headers):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]

    for row_data in data:
        row_values = [row_data.get(header, '') for header in headers]  # Преобразование словаря в список
        sheet.append(row_values)

    workbook.save(file_name)



if __name__ == '__main__':
    #file_path = "/Users/docha/PycharmProjects/Tools_for_buh/Kres/test/_svod_2312.xlsx"
    #file_path = "/Users/docha/Library/CloudStorage/GoogleDrive-kres.auto79@gmail.com/Мой диск/2023-12/avansiaruanned/_svod_2312.xlsx"
    file_path = "/Users/docha/Library/CloudStorage/GoogleDrive-kres.auto79@gmail.com/Мой диск/2023-12/kaart/_svod_2312_1.xlsx"
    sheet_name = "svod"
    new_sheet_name = "zagruzka"
    kaibemaks = 0.2
    expense_account_prefix = '4'

    headers = [
        'Вид бухгалтерской операции',
        'Счет',
        'Контрагент НСО',
        'Дата документа',
        'Вид документа НСО',
        'Серия документа',
        'Номер документа',
        'Валюта отчета',
        'Курс валюты отчета',
        'Сумма',
        'Ставка НСО',
        'Сумма НСО',
        'Сумма с НСО',
        'Счет проводки НСО',
        'Применение НСО',
        'Аналитика по НСО',
        'Счет 50%',
        'Субконто1',
        'Субконто2',
        'Субконто3',
        'Субконто4',
        'Субконто5',
        'Субконто1 50%',
        'Субконто2 50%',
        'Субконто3 50%',
        'Субконто4 50%',
        'Субконто5 50%',
        'Основание',
        'Комментарий',
    ]  # Все 29 заголовков столбцов на новом листе (1c)

    kaibemaksuta_desc = [
        'pant',
    ]

    column_mappings = {
        'Контрагент НСО': 'reg nr',
        'Дата документа': 'kuupaev',
        'Номер документа': 'arve nr',
    } # Соответствия заголовков столбцов 1С и их имен в исходном файле

    # Общие настройки для row_settings
    base_row_settings = {
        'doc_type': 'EKA чек',
        'account_transaction_nso': '220.02',
        'currency': 'EUR',
        'currency_rate': 1
    }

    # Дополнительные настройки для recalculate_tax_and_total = True, чеки в Эстонии
    additional_settings_true = {
        'application_nso': '225.01 НСО',
        'analytics_nso': 'Eestis kauba soetamine'
    }

    # Дополнительные настройки для recalculate_tax_and_total = False, чеки не в Эстонии
    additional_settings_false = {
        'application_nso': '132.02 Предоплата по НСО',
        'analytics_nso': ''
    }

    # Выбор нужных настроек в зависимости от значения recalculate_tax_and_total
    recalculate_tax_and_total = False
    additional_settings = additional_settings_false if not recalculate_tax_and_total else additional_settings_true

    # Объединение общих настроек с дополнительными
    row_settings = {**base_row_settings, **additional_settings}

    data = read_excel(file_path, sheet_name)

    create_new_sheet_with_headers(file_path, new_sheet_name, headers)

    processed_data = process_sheet_data_for_new_sheet(data, column_mappings,headers, kaibemaks, kaibemaksuta_desc,
                                                      recalculate_tax_and_total)
    
    write_data_to_sheet(file_path, new_sheet_name, processed_data, headers)

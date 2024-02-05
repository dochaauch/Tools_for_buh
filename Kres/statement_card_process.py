import openpyxl
import re

def read_excel(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name, data_only=True)
    sheet = workbook[sheet_name]

    # Чтение заголовков из первой строки
    headers = [cell.value for cell in sheet[1]]

    # Чтение данных, начиная с второй строки
    svod_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: cell for i, cell in enumerate(row)}
        svod_data.append(row_data)

    return svod_data

def filter_dicts_by_key_start(data, start_string, column_name='Selgitus'):
    return list(filter(lambda d: str(d.get(column_name, '')).startswith(start_string), data))

def parse_selgitus(selgitus):
    parts = re.split(r'\s+', selgitus, maxsplit=2)
    return {
        'card_number': parts[0],
        'date': parts[1],
        'description': parts[2] if len(parts) > 2 else ''
    }


def extract_store_name(description):
    # Функция для извлечения названия магазина из описания
    # Разделение строки по обратному слэшу и возврат первой части
    return description.split('\\')[0]


def parse_transaction_description(description):
    # Регулярное выражение для извлечения необходимых частей
    #'47.00 GEL VAHETUSKURSS: 2.829621, KONVERTEERIMISTASU 0.17 EUR MagtiCom N20 Office 0192 Tbilisi'
    pattern = r'(\d+.\d+) (\w+) VAHETUSKURSS: (\d+.\d+), KONVERTEERIMISTASU (\d+.\d+) (\w+) (.+)'
    match = re.match(pattern, description)
    if match:
        amount, currency, exchange_rate, conversion_fee, conversion_currency, store_name = match.groups()
        return {
            "amount": amount,
            "currency": currency,
            "exchange_rate": exchange_rate,
            "conversion_fee": conversion_fee,
            "conversion_currency": conversion_currency,
            "store_name": store_name
        }
    else:
        return {"error": "Format does not match"}


def copy_data_to_new_sheet(workbook, source_sheet_name, target_sheet_name):
    # Удаляем существующий лист, если он уже есть
    if target_sheet_name in workbook.sheetnames:
        del workbook[target_sheet_name]
    source_sheet = workbook[source_sheet_name]
    if target_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(target_sheet_name)
    target_sheet = workbook[target_sheet_name]

    for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, max_col=source_sheet.max_column):
        for cell in row:
            target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

def add_new_columns(sheet, column_headers):
    for i, header in enumerate(column_headers, start=sheet.max_column + 1):
        sheet.cell(row=1, column=i, value=header)


def find_column_indexes(sheet, column_headers):
    header_row = sheet[1]  # Предполагается, что заголовки находятся в первой строке
    header_to_index = {}
    for cell in header_row:
        if cell.value in column_headers:
            header_to_index[cell.value] = cell.column
    return header_to_index


def process_and_fill_data(sheet, column_headers, start_string):
    column_indexes = find_column_indexes(sheet, column_headers)

    for row in range(2, sheet.max_row + 1):
        selgitus_cell_value = sheet.cell(row=row, column=12).value  # Assuming Selgitus is in the 12 column
        selgitus = str(selgitus_cell_value) if selgitus_cell_value is not None else ''
        if selgitus.startswith(start_string):
            parsed = parse_selgitus(selgitus)
            store_data = parse_transaction_description(parsed['description'])

            # Изменённое условие для заполнения столбца "Магазин"
            if not parsed['description'][0].isdigit() :
                sheet.cell(row=row, column=column_indexes["Магазин"],
                           value=extract_store_name(parsed['description']))

            if 'error' not in store_data:
                # Обработка и преобразование суммы и конвертационной платы в числа
                try:
                    amount = float(store_data['amount'])
                    conversion_fee = float(store_data['conversion_fee'])
                    currency_rate = float(store_data['exchange_rate'])
                except ValueError:
                    # В случае, если преобразование не удаётся, можно записать 0 или оставить пустым
                    amount = 0
                    conversion_fee = 0
                sheet.cell(row=row, column=column_indexes["Магазин"],
                           value=extract_store_name(store_data['store_name']))
                sheet.cell(row=row, column=column_indexes["Сумма_вал"], value=amount)
                sheet.cell(row=row, column=column_indexes["Валюта"], value=store_data['currency'])
                sheet.cell(row=row, column=column_indexes["Конвертация"], value=conversion_fee)
                sheet.cell(row=row, column=column_indexes["Курс"], value=currency_rate)



def main():
    file_path = "/Users/docha/Library/CloudStorage/GoogleDrive-kres.auto79@gmail.com/Мой диск/pank/2023_statement.xlsx"
    original_sheet_name = "2023_statement"
    processed_sheet_name = "processed"
    new_columns = ["Магазин", "Сумма_вал", "Валюта", "Конвертация", "Курс"]
    start_string = '547372**'  # строка, с которой должны начинаться значения

    workbook = openpyxl.load_workbook(file_path)
    copy_data_to_new_sheet(workbook, original_sheet_name, processed_sheet_name)
    add_new_columns(workbook[processed_sheet_name], new_columns)
    process_and_fill_data(workbook[processed_sheet_name], new_columns, start_string)
    workbook.save(file_path)


if __name__ == '__main__':
    main()
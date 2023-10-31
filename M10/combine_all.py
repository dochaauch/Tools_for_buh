import glob
import os
import shutil

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def combine_all_data(folder_path):
    # Чтение всех файлов CSV и XLSX, начинающихся с 'combine' в указанной папке
    csv_files = glob.glob(folder_path + '/combine*.csv')
    xlsx_files = glob.glob(folder_path + '/combine*.xlsx')

    # Обработка CSV файлов
    combined_csv = combine_csv(csv_files)

    # Обработка XLSX файлов
    combined_xlsx = combine_xlsx(xlsx_files)

    # Сохранение объединенных данных в новых файлах CSV и XLSX
    csv_save_path = folder_path + '/combine_total.csv'
    xlsx_save_path = folder_path + '/combine_total.xlsx'
    combined_csv.to_csv(csv_save_path, index=False)
    combined_xlsx.save(xlsx_save_path)

    move_files_to_done(csv_files + xlsx_files, folder_path)


def combine_csv(csv_files):
    # Создание пустого DataFrame для объединения данных CSV
    combined_csv = pd.DataFrame()

    # Чтение и объединение файлов CSV
    for file in csv_files:
        data = pd.read_csv(file)
        combined_csv = combined_csv._append(data, ignore_index=True)

    # Преобразование 'kuupaev' в формат даты
    combined_csv['kuupaev'] = pd.to_datetime(combined_csv['kuupaev'], format='%d.%m.%y', errors='coerce')

    # Сортировка данных по столбцу 'kuupaev'
    combined_csv = combined_csv.sort_values(by='kuupaev')

    # Преобразование 'kuupaev' обратно в исходный формат
    combined_csv['kuupaev'] = combined_csv['kuupaev'].dt.strftime('%d.%m.%y')

    # Удаление первого столбца (индекса)
    combined_csv = combined_csv.iloc[:, 1:]

    return combined_csv


def combine_xlsx(xlsx_files):
    # Создание пустого Workbook для объединения данных XLSX
    combined_xlsx = Workbook()

    # Чтение и объединение файлов XLSX
    for file in xlsx_files:
        wb = load_workbook(file, read_only=True)
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            if sheetname not in combined_xlsx.sheetnames:
                combined_xlsx.create_sheet(title=sheetname)
            combined_ws = combined_xlsx[sheetname]
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i == 1 and combined_ws.max_row == 1:  # Если это первая строка и в объединенном листе еще нет заголовков
                    combined_ws.append(row)
                elif i > 1:  # Пропускаем первую строку, чтобы избежать дублирования заголовков
                    combined_ws.append(row)

    return combined_xlsx


def move_files_to_done(all_files, folder_path):
    # Создание папки "done", если она не существует
    done_folder = folder_path + '/done'
    os.makedirs(done_folder, exist_ok=True)

    # Перемещение обработанных файлов в папку "done"
    for file in all_files:
        file_name = os.path.basename(file)  # Извлекаем только имя файла
        if file_name != 'combine_total.csv' and file_name != 'combine_total.xlsx':
            destination = os.path.join(done_folder, file_name)  # Полный путь к файлу назначения
            shutil.move(file, destination)


def main():
    # Путь к папке с файлами CSV и XLSX
    folder_path = '/Users/docha/Library/CloudStorage/GoogleDrive-mob37256213753@gmail.com/Мой диск/Metsa10/2022/avans  '
    combine_all_data(folder_path)


if __name__ == '__main__':
    main()
